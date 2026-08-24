import json
import re
import openpyxl
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill
from io import BytesIO
from typing import List, Dict, Any, Tuple, Optional
from app.models.schemas import SheetInfo, CellStyle, CellFont, CellFill, CellAlignment, CellBorders, CellBorder


def _sanitize_col(original: str, seen: Dict[str, str]) -> str:
    """
    Strip everything except letters and digits, collapse runs to a single
    underscore, trim leading/trailing underscores.
    Guarantees uniqueness within `seen` (maps sanitized -> original).
    """
    safe = re.sub(r'[^a-zA-Z0-9]+', '_', original).strip('_') or 'col'
    candidate = safe
    counter = 2
    while candidate in seen:          # collision — another col already got this name
        candidate = f"{safe}_{counter}"
        counter += 1
    return candidate

def normalise_color(color) -> str:
    if not color: return None
    if hasattr(color, 'rgb') and isinstance(color.rgb, str):
        # ARGB to RGB
        rgb = color.rgb
        if len(rgb) == 8: return rgb[2:]
        return rgb
    return None

def extract_cell_style(cell) -> Optional[CellStyle]:
    style = CellStyle()
    has_any = False

    # Font
    if cell.font:
        f = CellFont()
        if cell.font.b: f.bold = True; has_any = True
        if cell.font.i: f.italic = True; has_any = True
        if cell.font.u: f.underline = True; has_any = True
        if cell.font.strike: f.strikethrough = True; has_any = True
        if cell.font.sz: f.size = cell.font.sz; has_any = True
        color = normalise_color(cell.font.color)
        if color: f.color = color; has_any = True
        if cell.font.name: f.name = cell.font.name; has_any = True
        if has_any: style.font = f

    # Fill
    if isinstance(cell.fill, PatternFill) and cell.fill.patternType == 'solid':
        color = normalise_color(cell.fill.fgColor)
        if color:
            style.fill = CellFill(bgColor=color)
            has_any = True

    # Alignment
    if cell.alignment:
        a = CellAlignment()
        if cell.alignment.horizontal: a.horizontal = cell.alignment.horizontal; has_any = True
        if cell.alignment.vertical: a.vertical = cell.alignment.vertical; has_any = True
        if cell.alignment.wrapText: a.wrapText = True; has_any = True
        if has_any: style.alignment = a

    # Borders
    if cell.border:
        b = CellBorders()
        has_border = False
        for side_name in ['top', 'bottom', 'left', 'right']:
            side = getattr(cell.border, side_name)
            if side and side.style:
                color = normalise_color(side.color)
                setattr(b, side_name, CellBorder(style=side.style, color=color))
                has_border = True
        if has_border:
            style.borders = b
            has_any = True

    if cell.number_format and cell.number_format != 'General':
        style.numberFormat = cell.number_format
        has_any = True

    return style if has_any else None

import os as _os
_PROJECT_ROOT = _os.path.dirname(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
COLUMN_MAP_PATH = _os.path.join(_PROJECT_ROOT, "column_maps.json")


def _save_column_map(sheet_name: str, column_map: Dict[str, str]) -> None:
    """Append/update this sheet's column map in the local JSON file."""
    try:
        with open(COLUMN_MAP_PATH, "r") as f:
            all_maps = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        all_maps = {}

    all_maps[sheet_name] = column_map

    with open(COLUMN_MAP_PATH, "w") as f:
        json.dump(all_maps, f, indent=2)


def add_column_to_map(col_name: str, original_name: str) -> None:
    """Add a new column mapping across all sheets in column_maps.json."""
    try:
        with open(COLUMN_MAP_PATH, "r") as f:
            all_maps = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        all_maps = {}

    for sheet in all_maps:
        all_maps[sheet][col_name] = original_name

    with open(COLUMN_MAP_PATH, "w") as f:
        json.dump(all_maps, f, indent=2)


def parse_excel(file_content: bytes) -> List[SheetInfo]:
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Simple table detection: find first row with data as header
        rows = list(ws.rows)
        if not rows: continue
        
        header_row_idx = -1
        col_mapping: Dict[int, str] = {}  # c_idx -> original name (sanitized after detection)

        for i, row in enumerate(rows):
            # Check for headers
            row_vals = []
            row_map = {}
            for c_idx, cell in enumerate(row):
                val = str(cell.value).strip() if cell.value is not None else ""
                if val:
                    row_map[c_idx] = val
                    row_vals.append(val)
            
            if len(row_vals) >= 2:
                header_row_idx = i
                col_mapping = row_map  # c_idx -> original name; sanitized below
                break
        
        if header_row_idx == -1:
            results.append(SheetInfo(
                data=[], titleRows=[], columns=[], columnMap={}, styles={},
                headerOffset=0, sheetName=sheet_name
            ))
            continue

        # -----------------------------------------------------------------
        # Build sanitized col_mapping: cell index -> sanitized name
        # and column_map:             sanitized name -> original name
        # col_mapping still maps c_idx -> original at this point; sanitize now.
        # -----------------------------------------------------------------
        column_map: Dict[str, str] = {}   # sanitized -> original
        sanitized_col_mapping: Dict[int, str] = {}  # c_idx -> sanitized name

        sorted_indices = sorted(col_mapping.keys())
        for c_idx in sorted_indices:
            original = col_mapping[c_idx]
            safe = _sanitize_col(original, column_map)
            column_map[safe] = original
            sanitized_col_mapping[c_idx] = safe

        columns = [sanitized_col_mapping[idx] for idx in sorted_indices]

        data = []
        styles = {}

        # Helper to extract row data using sanitized column names
        def extract_row_data(row, target_idx, is_title=False):
            row_id = f"title_{target_idx}" if is_title else target_idx + 1
            row_dict = {"ID": row_id}
            has_val = False

            for c_idx, safe_name in sanitized_col_mapping.items():
                if c_idx < len(row):
                    cell = row[c_idx]
                    val = cell.value
                    row_dict[safe_name] = val if val is not None else ""
                    if val is not None:
                        has_val = True

                    style = extract_cell_style(cell)
                    if style:
                        style_row_idx = (-1000 + target_idx) if is_title else target_idx
                        styles[f"{style_row_idx}:{safe_name}"] = style
            return row_dict if has_val else None

        # Parse data rows
        for r_idx, row in enumerate(rows[header_row_idx + 1:], start=0):
            row_data = extract_row_data(row, r_idx, is_title=False)
            if row_data:
                data.append(row_data)

        # Parse header styles (use sanitized names as style keys)
        header_row = rows[header_row_idx]
        for c_idx, safe_name in sanitized_col_mapping.items():
            if c_idx < len(header_row):
                cell = header_row[c_idx]
                style = extract_cell_style(cell)
                if style:
                    styles[f"-1:{safe_name}"] = style

        # Parse title rows (rows before header)
        title_rows = []
        for r_idx, row in enumerate(rows[:header_row_idx]):
            row_data = extract_row_data(row, r_idx, is_title=True)
            if row_data:
                title_rows.append(row_data)

        _save_column_map(sheet_name, column_map)

        results.append(SheetInfo(
            data=data,
            titleRows=title_rows,
            columns=columns,
            columnMap=column_map,
            styles=styles,
            headerOffset=header_row_idx,
            sheetName=sheet_name
        ))

    return results